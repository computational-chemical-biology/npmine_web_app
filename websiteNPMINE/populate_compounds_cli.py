import os
import random
import time
from datetime import datetime

import click
from flask import current_app
from openpyxl import load_workbook
import pubchempy as pcp
import requests

from websiteNPMINE import db
from websiteNPMINE.compounds.utils import save_compound_image
from websiteNPMINE.models import Accounts, Compounds, DOI, Taxa


DEFAULT_COMPOUNDS_FILE = 'dados_teste/Table S1_total_unique_inchikeys.xlsx'
DEFAULT_TAXA_FILE = 'dados_teste/df_taxons_final.xlsx'


def register_commands(app):
    app.cli.add_command(import_compounds_command)
    app.cli.add_command(import_taxa_command)


def _resolve_project_path(relative_or_absolute_path):
    if os.path.isabs(relative_or_absolute_path):
        return relative_or_absolute_path

    project_root = os.path.abspath(os.path.join(current_app.root_path, os.pardir))
    return os.path.join(project_root, relative_or_absolute_path)


def _require_admin_user():
    admin = Accounts.query.filter_by(username='admin').first()
    if admin:
        return admin

    raise click.ClickException(
        "Admin account not found. Run 'flask db upgrade' first so bootstrap migrations can seed it."
    )


def _get_or_create_doi(article_url):
    if not article_url:
        return None

    doi = DOI.query.filter_by(doi=article_url).first()
    if doi:
        return doi

    doi = DOI(doi=article_url)
    db.session.add(doi)
    db.session.flush()
    return doi


def _stringify_classifier_field(value):
    if isinstance(value, list):
        return ', '.join(str(item) for item in value)
    return value


def _fetch_pubchem_metadata(inchi_key, retries=3):
    if not inchi_key:
        return {}

    while retries > 0:
        try:
            time.sleep(random.uniform(1.0, 2.0))
            compounds = pcp.get_compounds(inchi_key, 'inchikey')
            if not compounds:
                return {}

            compound = compounds[0]
            return {
                'smiles': compound.isomeric_smiles or compound.canonical_smiles,
                'compound_name': compound.iupac_name,
                'pubchem_id': compound.cid,
            }
        except Exception as exc:
            retries -= 1
            if retries == 0:
                current_app.logger.warning("PubChem lookup failed for %s: %s", inchi_key, exc)
                return {}
            time.sleep(5)


def _fetch_npclassifier_metadata(smiles):
    if not smiles:
        return {}

    try:
        response = requests.get(
            'https://npclassifier.gnps2.org/classify',
            params={'smiles': smiles},
            timeout=20,
        )
        response.raise_for_status()
        payload = response.json()
        return {
            'class_results': _stringify_classifier_field(payload.get('class_results')),
            'superclass_results': _stringify_classifier_field(payload.get('superclass_results')),
            'pathway_results': _stringify_classifier_field(payload.get('pathway_results')),
            'isglycoside': payload.get('isglycoside'),
        }
    except Exception as exc:
        current_app.logger.warning("NPClassifier lookup failed for smiles %s: %s", smiles, exc)
        return {}


def _find_existing_compound(article_url, inchi_key, smiles):
    query = Compounds.query.filter_by(article_url=article_url)
    if inchi_key:
        return query.filter_by(inchi_key=inchi_key).first()
    if smiles:
        return query.filter_by(smiles=smiles).first()
    return None


def _find_existing_taxon(article_url, verbatim, taxonid):
    query = Taxa.query.filter_by(article_url=article_url, verbatim=verbatim)
    if taxonid is None:
        return query.filter(Taxa.taxonid.is_(None)).first()
    return query.filter_by(taxonid=taxonid).first()


@click.command('import-compounds')
@click.option('--file', 'file_path', default=DEFAULT_COMPOUNDS_FILE, show_default=True, help='XLSX file with compounds.')
@click.option('--limit', type=int, default=None, help='Import at most this many spreadsheet rows.')
@click.option('--commit-every', type=int, default=100, show_default=True, help='Commit every N rows.')
@click.option('--skip-existing/--no-skip-existing', default=True, show_default=True, help='Skip compounds already imported for the same DOI/article.')
@click.option('--with-pubchem/--no-with-pubchem', default=False, show_default=True, help='Enrich missing data using PubChem.')
@click.option('--with-npclassifier/--no-with-npclassifier', default=False, show_default=True, help='Fetch NPClassifier metadata.')
@click.option('--with-images/--no-with-images', default=True, show_default=True, help='Generate compound images locally with RDKit.')
def import_compounds_command(file_path, limit, commit_every, skip_existing, with_pubchem, with_npclassifier, with_images):
    """Import compounds from an XLSX file."""
    admin = _require_admin_user()
    resolved_path = _resolve_project_path(file_path)

    if not os.path.exists(resolved_path):
        raise click.ClickException(f"Compounds file not found: {resolved_path}")

    workbook = load_workbook(resolved_path, read_only=True, data_only=True)
    sheet = workbook.active
    stats = {'created': 0, 'updated': 0, 'linked': 0, 'skipped': 0, 'failed': 0}

    click.echo(f"Importing compounds from {resolved_path}")

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if limit and (stats['created'] + stats['updated'] + stats['skipped'] + stats['failed']) >= limit:
            break

        if not row or len(row) < 8:
            stats['skipped'] += 1
            continue

        try:
            journal, smiles, article_url, inchi_key, exact_molecular_weight, pubchem_id, inchi, source = row[:8]
            doi = _get_or_create_doi(article_url)

            compound = _find_existing_compound(article_url, inchi_key, smiles) if skip_existing else None
            if compound:
                stats['skipped'] += 1
            else:
                status = 'public'
                compound_data = {
                    'journal': journal,
                    'smiles': smiles,
                    'article_url': article_url,
                    'inchi_key': inchi_key,
                    'exact_molecular_weight': float(exact_molecular_weight) if exact_molecular_weight else None,
                    'pubchem_id': str(pubchem_id) if pubchem_id else None,
                    'inchi': inchi,
                    'source': source,
                    'user_id': admin.id,
                    'created_at': datetime.utcnow(),
                    'updated_at': datetime.utcnow(),
                    'status': status,
                    'ispublic': status == 'public',
                    'compound_name': None,
                }

                if with_pubchem:
                    pubchem_data = _fetch_pubchem_metadata(inchi_key)
                    compound_data['smiles'] = compound_data['smiles'] or pubchem_data.get('smiles')
                    compound_data['compound_name'] = pubchem_data.get('compound_name')
                    compound_data['pubchem_id'] = compound_data['pubchem_id'] or (
                        str(pubchem_data['pubchem_id']) if pubchem_data.get('pubchem_id') else None
                    )

                if with_npclassifier and compound_data['smiles']:
                    compound_data.update(_fetch_npclassifier_metadata(compound_data['smiles']))

                compound = Compounds(**compound_data)
                db.session.add(compound)
                db.session.flush()
                stats['created'] += 1

            if doi and doi not in compound.dois:
                compound.dois.append(doi)
                stats['linked'] += 1

            if with_images and compound.smiles and not compound.compound_image:
                compound.compound_image = save_compound_image(compound.id, compound.smiles)
                stats['updated'] += 1

            if row_number % commit_every == 0:
                db.session.commit()
                click.echo(
                    f"Processed row {row_number}: created={stats['created']} skipped={stats['skipped']} failed={stats['failed']}"
                )
        except Exception as exc:
            db.session.rollback()
            stats['failed'] += 1
            click.echo(f"Row {row_number} failed: {exc}", err=True)

    db.session.commit()
    workbook.close()

    click.echo(
        "Compound import complete: "
        f"created={stats['created']} updated={stats['updated']} linked={stats['linked']} "
        f"skipped={stats['skipped']} failed={stats['failed']}"
    )


@click.command('import-taxa')
@click.option('--file', 'file_path', default=DEFAULT_TAXA_FILE, show_default=True, help='XLSX file with taxa.')
@click.option('--limit', type=int, default=None, help='Import at most this many spreadsheet rows.')
@click.option('--commit-every', type=int, default=100, show_default=True, help='Commit every N rows.')
@click.option('--skip-existing/--no-skip-existing', default=True, show_default=True, help='Skip taxa already imported for the same DOI/article.')
def import_taxa_command(file_path, limit, commit_every, skip_existing):
    """Import taxa from an XLSX file."""
    admin = _require_admin_user()
    resolved_path = _resolve_project_path(file_path)

    if not os.path.exists(resolved_path):
        raise click.ClickException(f"Taxa file not found: {resolved_path}")

    workbook = load_workbook(resolved_path, read_only=True, data_only=True)
    sheet = workbook.active
    stats = {'created': 0, 'linked': 0, 'skipped': 0, 'failed': 0}

    click.echo(f"Importing taxa from {resolved_path}")

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if limit and (stats['created'] + stats['skipped'] + stats['failed']) >= limit:
            break

        if not row or len(row) < 8:
            stats['skipped'] += 1
            continue

        try:
            article_url, verbatim, odds, datasourceid, taxonid, classificationpath, classificationrank, matchtype = row[:8]
            doi = _get_or_create_doi(article_url)

            taxon = _find_existing_taxon(article_url, verbatim, taxonid) if skip_existing else None
            if taxon:
                stats['skipped'] += 1
            else:
                taxon = Taxa(
                    article_url=article_url,
                    verbatim=verbatim,
                    odds=float(odds) if odds is not None else None,
                    datasourceid=datasourceid,
                    taxonid=int(taxonid) if taxonid is not None else None,
                    classificationpath=classificationpath,
                    classificationrank=classificationrank,
                    matchtype=matchtype,
                    user_id=admin.id,
                    created_at=datetime.utcnow(),
                )
                db.session.add(taxon)
                db.session.flush()
                stats['created'] += 1

            if doi and doi not in taxon.dois:
                taxon.dois.append(doi)
                stats['linked'] += 1

            if row_number % commit_every == 0:
                db.session.commit()
                click.echo(
                    f"Processed row {row_number}: created={stats['created']} skipped={stats['skipped']} failed={stats['failed']}"
                )
        except Exception as exc:
            db.session.rollback()
            stats['failed'] += 1
            click.echo(f"Row {row_number} failed: {exc}", err=True)

    db.session.commit()
    workbook.close()

    click.echo(
        "Taxa import complete: "
        f"created={stats['created']} linked={stats['linked']} skipped={stats['skipped']} failed={stats['failed']}"
    )
