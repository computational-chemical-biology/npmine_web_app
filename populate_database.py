import psycopg2
from openpyxl import load_workbook
from tqdm import tqdm
import bcrypt
from datetime import datetime
import pubchempy as pcp
from rdkit import Chem
from rdkit.Chem import Draw
import os
import requests
import json
from dotenv import load_dotenv

load_dotenv()

# Retrieve environment variables
DB_HOST = os.getenv('DB_HOST')
DB_PORT = os.getenv('DB_PORT')
DB_NAME = os.getenv('DB_NAME')
DB_USER = os.getenv('DB_USER')
DB_PASSWORD = os.getenv('DB_PASSWORD')
NPMINE_WEB_APP_PASSWORD = os.getenv('NPMINE_WEB_APP_PASSWORD')
NPMINE_WEB_APP_EMAIL = os.getenv('NPMINE_WEB_APP_EMAIL')

# Establish a connection to the database using environment variables
conn = psycopg2.connect(
    host=DB_HOST,
    port=DB_PORT,
    database=DB_NAME,
    user=DB_USER,
    password=DB_PASSWORD
)

# Create a cursor object to execute SQL queries
cursor = conn.cursor()

# Define the get_or_create function for DOI
def get_or_create_doi(article_url):
    cursor.execute("SELECT id FROM doi WHERE doi = %s", (article_url,))
    doi_id = cursor.fetchone()
    if not doi_id:
        cursor.execute("INSERT INTO doi (doi) VALUES (%s) RETURNING id", (article_url,))
        doi_id = cursor.fetchone()[0]
    else:
        doi_id = doi_id[0]
    return doi_id

# Define the admin role ID, editor role ID, and user role ID
admin_role_id = 1
editor_role_id = 2
user_role_id = 3  # Add this line

# Insert roles if not already present
cursor.execute("SELECT id FROM role WHERE id = %s", (admin_role_id,))
if not cursor.fetchone():
    cursor.execute("INSERT INTO role (id, name) VALUES (%s, %s)", (admin_role_id, 'admin'))

cursor.execute("SELECT id FROM role WHERE id = %s", (editor_role_id,))
if not cursor.fetchone():
    cursor.execute("INSERT INTO role (id, name) VALUES (%s, %s)", (editor_role_id, 'editor'))

cursor.execute("SELECT id FROM role WHERE id = %s", (user_role_id,))  # Add this line
if not cursor.fetchone():  # Add this line
    cursor.execute("INSERT INTO role (id, name) VALUES (%s, %s)", (user_role_id, 'user'))  # Add this line

# Insert the "admin" user if not already present
cursor.execute("SELECT id FROM accounts WHERE username = %s", ('admin',))
admin_id = cursor.fetchone()
if not admin_id:
    # Hash the admin's password
    admin_password = NPMINE_WEB_APP_PASSWORD.encode('utf-8')
    salt = bcrypt.gensalt()
    admin_password_hash = bcrypt.hashpw(admin_password, salt).decode('utf-8')

    # Insert the admin user with the created_at field and admin role_id
    cursor.execute(
        "INSERT INTO accounts (username, email, password, created_at, role_id) VALUES (%s, %s, %s, %s, %s) RETURNING id",
        ('admin', NPMINE_WEB_APP_EMAIL, admin_password_hash, datetime.utcnow(), admin_role_id)
    )
    admin_id = cursor.fetchone()[0]
else:
    admin_id = admin_id[0]

# Function to generate compound images using RDKit
def generate_compound_image(smiles, compound_id):
    try:
        img_folder = 'websiteNPMINE/static/compound_images'  # Replace with your actual image folder path
        img_path = os.path.join(img_folder, f'{compound_id}.png')

        if not os.path.exists(img_path):
            mol = Chem.MolFromSmiles(smiles)
            img = Draw.MolToImage(mol, size=(200, 200))
            img.save(img_path)
        
        return f'compound_images/{compound_id}.png'
    except Exception as e:
        print(f"Error generating compound image: {e}")
        return None

# Populate the Compounds table from XLSX
workbook_compounds = load_workbook('dados_teste/Table S1_total_unique_inchikeys.xlsx')
sheet_compounds = workbook_compounds.active

# Dictionary to store article_url to doi_id mapping
article_url_to_doi_id = {}

# Set to store unique combinations of doi_id and compound_id
unique_doicomp_pairs = set()

total_compounds = sheet_compounds.max_row - 1  # Exclude the header row

with tqdm(total=total_compounds, desc='Populating Compounds') as pbar:
    for row in sheet_compounds.iter_rows(min_row=2, values_only=True):
        if len(row) < 8:
            continue

        journal, smiles, article_url, inchi_key, exact_molecular_weight, pubchem_id, inchi, source = row[:8]

        # Check if the article_url is already mapped to a doi_id
        if article_url in article_url_to_doi_id:
            doi_id = article_url_to_doi_id[article_url]
        else:
            doi_id = get_or_create_doi(article_url)
            article_url_to_doi_id[article_url] = doi_id

        # Handle empty values
        exact_molecular_weight = float(exact_molecular_weight) if exact_molecular_weight else None
        pubchem_id = int(pubchem_id) if pubchem_id else None

        # Use PubChemPy to fetch compound_name using InChI Key
        compound_name = None  # Initialize with None

        if inchi_key:
            try:
                compounds = pcp.get_compounds(inchi_key, 'inchikey')
                if compounds:
                    compound = compounds[0]
                    if not smiles and compound.isomeric_smiles:
                        smiles = compound.isomeric_smiles
                    if compound.iupac_name:
                        compound_name = compound.iupac_name  # Store the compound name
                else:
                    print(f"No data found for InChI Key: {inchi_key}")
            except Exception as e:
                print(f"Error fetching compound_name data: {e}")

        # Make a request to the NPclassifier API to get classification results
        api_url = f'https://npclassifier.gnps2.org/classify?smiles={smiles}'
        response = requests.get(api_url)
        try:
            api_data = response.json()
        except json.decoder.JSONDecodeError as e:
            print(f"Error decoding JSON response: {e}")
            api_data = None  # Or any other fallback value

        # Extract classification results from the API response
        class_results = None
        superclass_results = None
        pathway_results = None
        isglycoside = None

        if api_data is not None:
            class_results = api_data.get('class_results', None)
            superclass_results = api_data.get('superclass_results', None)
            pathway_results = api_data.get('pathway_results', None)
            isglycoside = api_data.get('isglycoside', None)
        else:
            print("API response is None, skipping classification results")

        # Insert the compound into the Compounds table
        cursor.execute(
            "INSERT INTO compounds (journal, smiles, article_url, inchi_key, exact_molecular_weight, pubchem_id, inchi, source, user_id, created_at, compound_name, status, class_results, superclass_results, pathway_results, isglycoside) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,  %s,  %s,  %s) RETURNING id",
            (journal, smiles, article_url, inchi_key, exact_molecular_weight, pubchem_id, inchi, source, admin_id, datetime.utcnow(), compound_name, 'private', class_results, superclass_results, pathway_results, isglycoside)
        )
        
        compound_id = cursor.fetchone()[0]

        # Check if the pair of doi_id and compound_id is unique
        if (doi_id, compound_id) not in unique_doicomp_pairs:
            # Insert the compound ID and DOI ID into the doicomp table
            cursor.execute("INSERT INTO doicomp (doi_id, compound_id) VALUES (%s, %s)", (doi_id, compound_id))
            unique_doicomp_pairs.add((doi_id, compound_id))

        # Generate and store the compound image
        if not smiles:
            continue

        compound_image_path = generate_compound_image(smiles, compound_id)
        if compound_image_path:
            cursor.execute(
                "UPDATE compounds SET compound_image = %s WHERE id = %s",
                (compound_image_path, compound_id)
            )

        # Update the progress bar
        pbar.update()

# Populate the Taxa table from XLSX
workbook_taxons = load_workbook('dados_teste/df_taxons_final.xlsx')
sheet_taxons = workbook_taxons.active

# Dictionary to store article_url to doi_id mapping
article_url_to_doi_id = {}

# Set to store unique combinations of doi_id and taxon_id
unique_doitaxa_pairs = set()

total_taxons = sheet_taxons.max_row - 1  # Exclude the header row

with tqdm(total=total_taxons, desc='Populating Taxa') as pbar:
    for row in sheet_taxons.iter_rows(min_row=2, values_only=True):
        if len(row) < 8:
            continue

        article_url, verbatim, odds, datasourceid, taxonid, classificationpath, classificationrank, matchtype = row[:8]

        # Check if the article_url is already mapped to a doi_id
        if article_url in article_url_to_doi_id:
            doi_id = article_url_to_doi_id[article_url]
        else:
            doi_id = get_or_create_doi(article_url)
            article_url_to_doi_id[article_url] = doi_id

        # Insert the taxon into the Taxa table
        cursor.execute(
            "INSERT INTO taxa (article_url, verbatim, odds, datasourceid, taxonid, classificationpath, classificationrank, matchtype, user_id, created_at) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id",
            (article_url, verbatim, odds, datasourceid, taxonid, classificationpath, classificationrank, matchtype, admin_id, datetime.utcnow())
        )
        taxon_id = cursor.fetchone()[0]

        # Check if the pair of doi_id and taxon_id is unique
        if (doi_id, taxon_id) not in unique_doitaxa_pairs:
            # Insert the taxon ID and DOI ID into the doitaxa table
            cursor.execute("INSERT INTO doitaxa (doi_id, taxon_id) VALUES (%s, %s)", (doi_id, taxon_id))
            unique_doitaxa_pairs.add((doi_id, taxon_id))

        # Update the progress bar
        pbar.update()

# Commit the changes to the database
conn.commit()

# Close the cursor and connection
cursor.close()
conn.close()
